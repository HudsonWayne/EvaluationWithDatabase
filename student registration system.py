from tkinter import *
from datetime import date
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

# ---------- COLORS & STYLES ----------
background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

# ---------- WINDOW SETUP ----------
root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

# ---------- Paths ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IMAGE_DIR = os.path.join(BASE_DIR, "images")
EXCEL_PATH = os.path.join(BASE_DIR, "student_data.xlsx")

# ---------- EXCEL FILE CHECK ----------
file = pathlib.Path(EXCEL_PATH)
if not file.exists():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "StudentData"
    sheet.append(["Registration No", "Name", "Class", "Gender", "DOB", "Date of Registration", 
                  "Religion", "Skill", "Father Name", "Mother Name", "Father's Occupation", "Mother's Occupation"])
    wb.save(EXCEL_PATH)





##############Registration No##############
#noe eachtime you have to eneter Registration No


def registration_no():
    file = openpyxl.load_workbook('student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    
    max_row_value= sheet.cell(row=row, column=1).value
    
    try:
        Registration.set(max_row_value+1)
        
    except:
        Registration.set("1")
        
#################Clear##################
def Clear():
    Name.set(' ')
    DOB.set(' ')
    Religion.set(' ')
    Skill.set(' ')
    F_Name.set(' ')
    M_Name.set(' ')
    F_Occupation.set(' ')
    M_Occupation.set(' ')
    Father_Occupation.set(' ')
    Mother_Occupation.set(' ')
    Class.set("Select Class")
    
    registration_no()
    
    saveButton.config(state = 'normal')
    img1= PhotoImage(file= "images/user.png")
    lbl.config(image=img1)
    lbl.image = img1
    
    img= " "
    
        

# ---------- Exit ----------
def Exit():
    root.destroy()

# ---------- Show Image ----------
def showimage():
    global photo2
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file",
                                          filetypes=[("Image Files", "*.jpg *.jpeg *.png")])
    if filename:
        img = Image.open(filename)
        resized_image = img.resize((180, 180))
        photo2 = ImageTk.PhotoImage(resized_image)
        lbl.config(image=photo2)
        lbl.image = photo2


############SAVE#################
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("error, Select Gender!")
    
    



# ---------- Gender Selection ----------
def selection():
    global gender
    value = radio.get()
    gender = "Male" if value == 1 else "Female"
    print(gender)

# ---------- Email Bar ----------
Label(root, text="Email: hudsonnbenhuraa@gmail.com", bg="#f0687c",
      anchor='e', fg='white', font='arial 12', height=2).pack(side=TOP, fill="x")

# ---------- HEADER ----------
header_frame = Frame(root, bg="#c36464", height=80)
header_frame.pack(side=TOP, fill="x")

title_frame = Frame(header_frame, bg="#c36464")
title_frame.pack(side=LEFT, padx=190)

Label(title_frame, text="STUDENT REGISTRATION", bg="#c36464", fg='white',
      font='arial 20 bold', height=2).pack()

search_frame = Frame(header_frame, bg="#c36464")
search_frame.pack(side=RIGHT, padx=20)

Search = StringVar()
search_entry = Entry(search_frame, textvariable=Search, width=20, bd=2, font="arial 16")
search_entry.pack(side=LEFT, padx=10, pady=20)

def focus_search():
    search_entry.focus_set()

# ---------- Load Images ----------
def load_image(path, size):
    try:
        img = Image.open(path)
        img = img.resize(size)
        return ImageTk.PhotoImage(img)
    except Exception as e:
        print(f"Image not found: {path}")
        return None

search_img = load_image(os.path.join(IMAGE_DIR, "search.png"), (30, 30))
layer_img = load_image(os.path.join(IMAGE_DIR, "layer.jpg"), (40, 40))

if search_img:
    Srch = Button(search_frame, text="Search", compound=LEFT, image=search_img,
                  bg='#68ddfa', font="arial 12 bold", padx=10,
                  command=focus_search)
    Srch.pack(side=LEFT, padx=5)

if layer_img:
    Update_button = Button(root, image=layer_img, bg="#c36464")
    Update_button.place(x=110, y=64)

# ---------- REGISTRATION & DATE ----------
Label(root, text="Registration No", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = StringVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()

today = date.today()
Date.set(today.strftime("%d/%m/%Y"))
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

# ---------- STUDENT DETAILS ----------
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of Birth:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)
Label(obj, text="Class:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
Entry(obj, textvariable=Name, width=20, font="arial 10").place(x=160, y=50)

DOB = StringVar()
Entry(obj, textvariable=DOB, width=20, font="arial 10").place(x=160, y=100)

radio = IntVar()
Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection).place(x=150, y=150)
Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection).place(x=200, y=150)

Class = Combobox(obj, values=[str(i) for i in range(1, 13)], font="Roboto", width=17, state="r")
Class.place(x=630, y=50)
Class.set("Select Class")

Religion = StringVar()
Entry(obj, textvariable=Religion, width=20, font="arial 10").place(x=630, y=100)

Skill = StringVar()
Entry(obj, textvariable=Skill, width=20, font="arial 10").place(x=630, y=150)

# ---------- PARENT DETAILS ----------
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=30)
Label(obj2, text="Father's Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=80)
Label(obj2, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=30)
Label(obj2, text="Mother's Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=80)

F_Name = StringVar()
Entry(obj2, textvariable=F_Name, width=20, font="arial 10").place(x=180, y=30)

F_Occupation = StringVar()
Entry(obj2, textvariable=F_Occupation, width=20, font="arial 10").place(x=180, y=80)

M_Name = StringVar()
Entry(obj2, textvariable=M_Name, width=20, font="arial 10").place(x=670, y=30)

M_Occupation = StringVar()
Entry(obj2, textvariable=M_Occupation, width=20, font="arial 10").place(x=670, y=80)

# ---------- IMAGE FRAME ----------
f = Frame(root, bd=2, width=200, height=200, relief=GROOVE, bg="black")
f.place(x=1000, y=150)

lbl = Label(f, bg="black")
lbl.pack()

# ---------- Load and show user.png ----------
try:
    user_path = os.path.join(IMAGE_DIR, "user.png")
    user_img_pil = Image.open(user_path).resize((180, 180))
    user_img = ImageTk.PhotoImage(user_img_pil)
    lbl.config(image=user_img)
    lbl.image = user_img
except Exception as e:
    print("Could not load user.png:", e)

# ---------- Buttons ----------
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)
Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen").place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)

# ---------- MAIN LOOP ----------
root.mainloop()
